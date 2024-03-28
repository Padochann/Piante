import openpyxl
import requests
import os

# Cartella di destinazione
cartella_immagini = "D:/informatica/images/"

# Aprire il file Excel
wb = openpyxl.load_workbook("D:/informatica/linkimmagini.xlsx")
foglio = wb.active

# Ciclo per ogni riga del foglio

i = 1
for riga in range(1, foglio.max_row + 1):
    # Ottenere il link dell'immagine
    link_immagine = foglio.cell(riga, 1).value

    # Generare il nome del file
    nome_file_str = str(i)  # Converte nome_file in stringa
    i += 1

    # Scaricare l'immagine
    try:
        response = requests.get(link_immagine)
        with open(os.path.join(cartella_immagini, nome_file_str + ".jpg"), "wb") as f:
            f.write(response.content)
    except Exception as e:
        print(f"Errore durante il download dell'immagine '{nome_file_str}': {e}")

    # Stampare un messaggio di successo
    print(f"Immagine '{nome_file_str}.jpg' scaricata correttamente!")
